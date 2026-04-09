from flask import Blueprint, request, jsonify, send_file
import pandas as pd
from io import BytesIO
from data_processor import load_data, compute_metrics
from genai_helper import (
    generate_query_plan,
    execute_query_plan,
    generate_rag_response,
)

api_bp = Blueprint('api', __name__)

@api_bp.route("/api/insight", methods=["POST"])
def generate_insight():
    body = request.get_json(silent=True) or {}
    context = body.get("context", {}) or {}

    page = str(context.get("page", "dashboard")).strip().lower()
    dept_filter = str(context.get("department", "All")).strip()
    bu_filter = str(context.get("business_unit", "All")).strip()
    start_date = context.get("start_date")
    end_date = context.get("end_date")

    df = load_data()

    if dept_filter and dept_filter != "All" and "Department" in df.columns:
        df = df[df["Department"] == dept_filter]
    if bu_filter and bu_filter != "All" and "Business Unit" in df.columns:
        df = df[df["Business Unit"] == bu_filter]

    if (start_date or end_date) and "Start Date" in df.columns and "End Date" in df.columns:
        if start_date:
            df = df[pd.to_datetime(df["Start Date"], errors="coerce") >= pd.to_datetime(start_date)]
        if end_date:
            df = df[pd.to_datetime(df["End Date"], errors="coerce") <= pd.to_datetime(end_date)]

    if df.empty:
        return jsonify({
            "insight": "No data found for the selected filters. Try clearing filters or adjusting the date range.",
            "context_used": {"page": page, "department": dept_filter, "business_unit": bu_filter, "start_date": start_date, "end_date": end_date},
        })

    metrics = compute_metrics(df)

    def _top_series(s: pd.Series, n=3, ascending=False):
        if s is None or len(s) == 0:
            return []
        s2 = s.dropna()
        if len(s2) == 0:
            return []
        s2 = s2.sort_values(ascending=ascending).head(n)
        return [(str(idx), float(val)) for idx, val in s2.items()]

    top_depts = _top_series(metrics.get("dept"), n=3, ascending=False)
    low_depts = _top_series(metrics.get("dept"), n=1, ascending=True)
    top_bus = _top_series(metrics.get("bu"), n=3, ascending=False)
    low_bus = _top_series(metrics.get("bu"), n=1, ascending=True)

    ttype = metrics.get("ttype")
    top_type = None
    if isinstance(ttype, pd.Series) and len(ttype) > 0:
        top_type = (str(ttype.index[0]), float(ttype.iloc[0]))

    emp_agg = df.groupby("Emp ID").agg(
        Total_Hours=("Overall Training Duration (Planned Hrs)", "sum"),
        Employee_Name=("Employee Name", "first") if "Employee Name" in df.columns else ("Emp ID", "first"),
        Department=("Department", "first") if "Department" in df.columns else ("Emp ID", "first"),
        Business_Unit=("Business Unit", "first") if "Business Unit" in df.columns else ("Emp ID", "first"),
    )
    emp_agg["Completed"] = emp_agg["Total_Hours"] >= 20
    remaining = (20 - emp_agg["Total_Hours"]).clip(lower=0)
    near = (
        emp_agg[~emp_agg["Completed"]]
        .assign(Remaining=remaining[~emp_agg["Completed"]])
        .sort_values(by="Remaining", ascending=True)
        .head(5)
    )

    lines = []
    lines.append("**Insight summary (based on current filters):**")
    lines.append(f"- **Learning coverage (20+ hrs)**: **{metrics.get('coverage', 0)}%** across **{metrics.get('total_emp', 0)}** employees")
    lines.append(f"- **Total session hours (unique sessions)**: **{metrics.get('avg_hours', 0)}**")
    if top_type:
        lines.append(f"- **Most common training type**: **{top_type[0]}** ({top_type[1]}%)")

    if top_depts:
        lines.append("- **Top departments by completion %**:")
        for name, val in top_depts:
            lines.append(f"  - {name}: {val:.1f}%")
    if low_depts:
        lines.append(f"- **Lowest department**: **{low_depts[0][0]}** ({low_depts[0][1]:.1f}%)")

    if top_bus:
        lines.append("- **Top business units by completion %**:")
        for name, val in top_bus:
            lines.append(f"  - {name}: {val:.1f}%")
    if low_bus:
        lines.append(f"- **Lowest business unit**: **{low_bus[0][0]}** ({low_bus[0][1]:.1f}%)")

    if len(near) > 0:
        lines.append("- **Employees closest to 20 hours (actionable follow-up)**:")
        for _, row in near.iterrows():
            nm = str(row.get("Employee_Name", "")).strip()
            eid = str(_)
            rem = float(row.get("Remaining", 0))
            lines.append(f"  - {nm} (Emp ID {eid}): ~{rem:.1f} hrs remaining")

    if page == "employees":
        top_emp = emp_agg.sort_values(by="Total_Hours", ascending=False).head(5)
        lines.append("- **Top employees by total hours**:")
        for eid, row in top_emp.iterrows():
            nm = str(row.get("Employee_Name", "")).strip()
            hrs = float(row.get("Total_Hours", 0))
            lines.append(f"  - {nm} (Emp ID {eid}): {hrs:.1f} hrs")

    if page == "analytics" and "Training Start Month" in df.columns:
        monthly_emp = (
            df.groupby(["Training Start Month", "Emp ID"])["Overall Training Duration (Planned Hrs)"]
            .sum()
            .reset_index()
        )
        monthly_emp["certified"] = monthly_emp["Overall Training Duration (Planned Hrs)"] >= 20
        trend_df = monthly_emp.groupby("Training Start Month").agg(
            total=("Emp ID", "count"),
            certified=("certified", "sum")
        ).reset_index()
        if len(trend_df) >= 2:
            last = trend_df.iloc[-1]
            prev = trend_df.iloc[-2]
            delta = int(last["certified"]) - int(prev["certified"])
            direction = "up" if delta >= 0 else "down"
            lines.append(f"- **Latest month certified change**: **{direction} {abs(delta)}** vs previous month")

    return jsonify({
        "insight": "\n".join(lines),
        "context_used": {"page": page, "department": dept_filter, "business_unit": bu_filter, "start_date": start_date, "end_date": end_date},
    })

@api_bp.route("/api/chat", methods=["POST"])
def chat():
    body = request.get_json(silent=True) or {}
    message = body.get("message", "").strip()

    if not message:
        return jsonify({"error": "No message provided."}), 400

    df = load_data()

    schema_dict = {col: str(dtype) for col, dtype in df.dtypes.items()}

    gemini_plan = generate_query_plan(message, schema_dict)

    if gemini_plan.get("intent") == "error":
        return jsonify({
            "query": message,
            "final_response": (
                gemini_plan.get("response_text")
                or "We encountered an issue interpreting your request. Please check your GEMINI_API_KEY."
            ),
            "data_result": [],
        }), 500

    data_result = execute_query_plan(df, gemini_plan)

    final_response = generate_rag_response(message, data_result, gemini_plan)

    return jsonify({
        "query": message,
        "gemini_plan": gemini_plan,
        "data_result": data_result,
        "final_response": final_response,
    })


@api_bp.route("/api/export/csv")
def export_csv():
    """Export analytics data as CSV"""
    try:
        df = load_data()
        
        # Apply filters from query parameters
        dept_filter = request.args.get('department', 'All')
        bu_filter = request.args.get('business_unit', 'All')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if start_date:
            df = df[pd.to_datetime(df['Start Date'], errors='coerce') >= pd.to_datetime(start_date)]
        if end_date:
            df = df[pd.to_datetime(df['End Date'], errors='coerce') <= pd.to_datetime(end_date)]
        
        if dept_filter != 'All':
            df = df[df['Department'] == dept_filter]
        if bu_filter != 'All':
            df = df[df['Business Unit'] == bu_filter]
        
        # Create CSV buffer
        output = BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'analytics_export.csv'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@api_bp.route("/api/export/excel")
def export_excel():
    """Export analytics category tables as Excel"""
    try:
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.data_source import AxDataSource, StrRef
        from openpyxl.chart.series import DataPoint
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        df = load_data()
        
        # Apply filters from query parameters
        dept_filter = request.args.get('department', 'All')
        bu_filter = request.args.get('business_unit', 'All')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if start_date:
            df = df[pd.to_datetime(df['Start Date'], errors='coerce') >= pd.to_datetime(start_date)]
        if end_date:
            df = df[pd.to_datetime(df['End Date'], errors='coerce') <= pd.to_datetime(end_date)]
        
        if dept_filter != 'All':
            df = df[df['Department'] == dept_filter]
        if bu_filter != 'All':
            df = df[df['Business Unit'] == bu_filter]

        metrics = compute_metrics(df)

        def build_category_df(categories, label_name):
            rows = []
            totals = {key: 0 for key in ['A', 'B', 'C', 'D', 'E', 'F', 'total']}

            for label, data in categories.items():
                row = {
                    label_name: str(label),
                    'A': int(data.get('A', 0)),
                    'B': int(data.get('B', 0)),
                    'C': int(data.get('C', 0)),
                    'D': int(data.get('D', 0)),
                    'E': int(data.get('E', 0)),
                    'F': int(data.get('F', 0)),
                    'Total': int(data.get('total', 0)),
                }
                rows.append(row)
                for key in totals:
                    totals[key] += int(data.get(key, 0))

            rows.append({
                label_name: 'Grand Total',
                'A': totals['A'],
                'B': totals['B'],
                'C': totals['C'],
                'D': totals['D'],
                'E': totals['E'],
                'F': totals['F'],
                'Total': totals['total'],
            })

            return pd.DataFrame(rows, columns=[label_name, 'A', 'B', 'C', 'D', 'E', 'F', 'Total'])

        def build_series_df(series_obj, label_name, value_name, sort_desc=True):
            if series_obj is None or len(series_obj) == 0:
                return pd.DataFrame(columns=[label_name, value_name])

            df_series = (
                series_obj.rename(value_name)
                .rename_axis(label_name)
                .reset_index()
            )
            if sort_desc:
                df_series = df_series.sort_values(value_name, ascending=False)
            if value_name.endswith('%'):
                df_series[value_name] = df_series[value_name].round(1)
            return df_series

        def build_overall_hours_df():
            totals = {key: 0 for key in ['A', 'B', 'C', 'D', 'E', 'F']}
            for data in metrics.get('dept_hours_categories', {}).values():
                for key in totals:
                    totals[key] += int(data.get(key, 0))
            return pd.DataFrame(
                [{'Category': key, 'Employees': value} for key, value in totals.items()],
                columns=['Category', 'Employees']
            )

        def add_bar_chart(
            target_ws,
            source_ws,
            title,
            anchor,
            min_col,
            max_col,
            min_row,
            max_row,
            cat_col,
            chart_type='bar',
            width=8.0,
            height=5.2,
        ):
            if max_row <= min_row:
                return
            chart = BarChart()
            chart.type = chart_type
            chart.style = 2
            chart.height = height
            chart.width = width
            chart.grouping = 'clustered'
            chart.overlap = 0
            chart.varyColors = True

            data = Reference(source_ws, min_col=min_col, max_col=max_col, min_row=min_row + 1, max_row=max_row)
            chart.add_data(data, titles_from_data=False, from_rows=False)

            # Force string category references so Excel renders clean axis labels.
            cat_col_letter = get_column_letter(cat_col)
            category_ref = f"'{source_ws.title}'!${cat_col_letter}${min_row + 1}:${cat_col_letter}${max_row}"
            for series in chart.series:
                series.cat = AxDataSource(strRef=StrRef(f=category_ref))

            chart.legend = None
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None
            chart.x_axis.title = None
            chart.y_axis.title = None
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showSerName = False
            chart.dataLabels.showLegendKey = False
            chart.dataLabels.showPercent = False
            chart.dataLabels.dLblPos = 'outEnd'
            chart.gapWidth = 90

            palette = [
                "4F81BD", "C0504D", "9BBB59", "8064A2", "4BACC6",
                "F79646", "2E75B6", "70AD47", "A5A5A5", "FFC000"
            ]
            for series in chart.series:
                point_count = max(0, max_row - min_row)
                series.dPt = []
                for point_idx in range(point_count):
                    point = DataPoint(idx=point_idx)
                    point.graphicalProperties.solidFill = palette[point_idx % len(palette)]
                    point.graphicalProperties.line.solidFill = palette[point_idx % len(palette)]
                    series.dPt.append(point)

            target_ws.add_chart(chart, anchor)
        
        # Create Excel buffer
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_df = pd.DataFrame([
                ['Department Filter', dept_filter],
                ['Business Unit Filter', bu_filter],
                ['Date Range', f"{start_date or 'N/A'} to {end_date or 'N/A'}"],
                ['Total Employees', metrics.get('total_emp', 0)],
                ['Learning Coverage (20+ hrs)', metrics.get('coverage', 0)],
                ['Completion Rate', metrics.get('completion_rate', 0)],
            ], columns=['Metric', 'Value'])

            dept_df = build_category_df(metrics.get('dept_hours_categories', {}), 'Department')
            bu_df = build_category_df(metrics.get('bu_hours_categories', {}), 'Business Unit')
            dept_completion_df = build_series_df(metrics.get('dept'), 'Department', 'Completion %')
            bu_completion_df = build_series_df(metrics.get('bu'), 'Business Unit', 'Completion %')
            training_type_df = build_series_df(metrics.get('ttype'), 'Training Type', 'Share %')
            training_type_20_df = build_series_df(metrics.get('ttype_20hrs'), 'Training Type', '20H+ Share %')
            overall_hours_df = build_overall_hours_df()

            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            dept_df.to_excel(writer, sheet_name='Department Categories', index=False)
            bu_df.to_excel(writer, sheet_name='BU Categories', index=False)
            dept_completion_df.to_excel(writer, sheet_name='Dept Completion', index=False)
            bu_completion_df.to_excel(writer, sheet_name='BU Completion', index=False)
            training_type_df.to_excel(writer, sheet_name='Training Type Share', index=False)
            training_type_20_df.to_excel(writer, sheet_name='20H+ Type Share', index=False)
            overall_hours_df.to_excel(writer, sheet_name='Hours Distribution', index=False)

            workbook = writer.book
            charts_ws = workbook.create_sheet(title='Charts')
            charts_ws.sheet_view.showGridLines = False
            charts_ws['A1'] = 'Analytics Visuals'
            charts_ws['A2'] = f"Department: {dept_filter}"
            charts_ws['A3'] = f"Business Unit: {bu_filter}"
            charts_ws['A4'] = f"Date Range: {start_date or 'N/A'} to {end_date or 'N/A'}"
            charts_ws['A1'].font = Font(bold=True, size=16)
            for cell in ['A2', 'A3', 'A4']:
                charts_ws[cell].font = Font(bold=True)

            chart_titles = {
                'A6': 'Department Completion %',
                'T6': 'Business Unit Completion %',
                'A23': 'Training Type Share',
                'T23': 'Training Type Share (20H+)',
                'A40': 'Overall Hours Category Distribution',
            }
            for cell, title in chart_titles.items():
                charts_ws[cell] = title
                charts_ws[cell].font = Font(bold=True, size=12)

            add_bar_chart(
                charts_ws,
                writer.sheets['Dept Completion'],
                'Department Completion %',
                'A7',
                2,
                2,
                1,
                writer.sheets['Dept Completion'].max_row,
                1,
                'bar',
            )
            add_bar_chart(
                charts_ws,
                writer.sheets['BU Completion'],
                'Business Unit Completion %',
                'T7',
                2,
                2,
                1,
                writer.sheets['BU Completion'].max_row,
                1,
                'bar',
            )
            add_bar_chart(
                charts_ws,
                writer.sheets['Training Type Share'],
                'Training Type Share',
                'A24',
                2,
                2,
                1,
                writer.sheets['Training Type Share'].max_row,
                1,
                'col',
            )
            add_bar_chart(
                charts_ws,
                writer.sheets['20H+ Type Share'],
                'Training Type Share (20H+)',
                'T24',
                2,
                2,
                1,
                writer.sheets['20H+ Type Share'].max_row,
                1,
                'col',
            )
            add_bar_chart(
                charts_ws,
                writer.sheets['Hours Distribution'],
                'Overall Hours Category Distribution',
                'A41',
                2,
                2,
                1,
                writer.sheets['Hours Distribution'].max_row,
                1,
                'col',
            )

            # Auto-adjust column widths for every exported sheet
            for worksheet in writer.sheets.values():
                for column_cells in worksheet.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            cell_length = len(str(cell.value)) if cell.value is not None else 0
                            if cell_length > max_length:
                                max_length = cell_length
                        except Exception:
                            pass
                    worksheet.column_dimensions[column_letter].width = max_length + 2
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'analytics_export.xlsx'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@api_bp.route("/api/export/pdf")
def export_pdf():
    """Export analytics data as PDF summary"""
    try:
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
            PageBreak,
            KeepTogether,
        )
        from reportlab.lib import colors
        from reportlab.graphics.shapes import Drawing, Rect, String, Line
        from datetime import datetime
        
        df = load_data()
        
        # Apply filters from query parameters
        dept_filter = request.args.get('department', 'All')
        bu_filter = request.args.get('business_unit', 'All')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if start_date:
            df = df[pd.to_datetime(df['Start Date'], errors='coerce') >= pd.to_datetime(start_date)]
        if end_date:
            df = df[pd.to_datetime(df['End Date'], errors='coerce') <= pd.to_datetime(end_date)]
        
        if dept_filter != 'All':
            df = df[df['Department'] == dept_filter]
        if bu_filter != 'All':
            df = df[df['Business Unit'] == bu_filter]
        
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name='Subtle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#475569'),
        ))
        styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            leading=16,
            textColor=colors.HexColor('#0f172a'),
            spaceAfter=0,
        ))
        styles.add(ParagraphStyle(
            name='CardText',
            parent=styles['BodyText'],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#334155'),
        ))
        styles.add(ParagraphStyle(
            name='CardLabel',
            parent=styles['BodyText'],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=2,
        ))
        styles.add(ParagraphStyle(
            name='CardValue',
            parent=styles['BodyText'],
            fontSize=18,
            leading=20,
            textColor=colors.HexColor('#334155'),
        ))

        metrics = compute_metrics(df)

        def _truncate(text, size=20):
            text = str(text)
            return text if len(text) <= size else f"{text[:size - 1]}..."

        def _series_pairs(series_obj, limit=None):
            if series_obj is None or len(series_obj) == 0:
                return []
            pairs = [(str(idx), float(val)) for idx, val in series_obj.items()]
            if limit is not None:
                pairs = pairs[:limit]
            return pairs

        def _nice_max(value, floor=10):
            if value <= 0:
                return floor
            if value <= 100:
                return 100 if floor == 100 else max(floor, int(((value + 9) // 10) * 10))
            magnitude = 10 ** (len(str(int(value))) - 1)
            return int(((value + magnitude - 1) // magnitude) * magnitude)

        def _hours_totals(categories):
            totals = {key: 0 for key in ['A', 'B', 'C', 'D', 'E', 'F']}
            for data in categories.values():
                for key in totals:
                    totals[key] += int(data.get(key, 0))
            return totals

        def build_highlights_panel():
            dept_pairs = sorted(_series_pairs(metrics.get('dept')), key=lambda item: item[1], reverse=True)
            bu_pairs = sorted(_series_pairs(metrics.get('bu')), key=lambda item: item[1], reverse=True)
            ttype_pairs = sorted(_series_pairs(metrics.get('ttype')), key=lambda item: item[1], reverse=True)
            ttype_20_pairs = sorted(_series_pairs(metrics.get('ttype_20hrs')), key=lambda item: item[1], reverse=True)

            highlight_items = [
                ("Top Department", f"{dept_pairs[0][0]} ({dept_pairs[0][1]:.1f}%)" if dept_pairs else "N/A"),
                ("Top Business Unit", f"{bu_pairs[0][0]} ({bu_pairs[0][1]:.1f}%)" if bu_pairs else "N/A"),
                ("Top Training Type", f"{ttype_pairs[0][0]} ({ttype_pairs[0][1]:.1f}%)" if ttype_pairs else "N/A"),
                ("Top 20H+ Type", f"{ttype_20_pairs[0][0]} ({ttype_20_pairs[0][1]:.1f}%)" if ttype_20_pairs else "N/A"),
                ("Training Records", str(len(df))),
                ("Avg Session Hours", str(metrics.get('avg_hours', 0))),
            ]

            cells = []
            for label_text, value_text in highlight_items:
                cells.append(Paragraph(
                    f"<font color='#64748b'>{label_text}</font><br/><b>{_truncate(value_text, 38)}</b>",
                    styles['CardText'],
                ))

            table = Table([cells[:3], cells[3:]], colWidths=[3.22 * inch] * 3)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#cbd5e1')),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            return table

        def build_metric_cards():
            card_items = [
                ("Total Employees", str(metrics.get('total_emp', 0)), colors.HexColor('#e0f2fe')),
                ("Learning Coverage", f"{metrics.get('coverage', 0)}%", colors.HexColor('#dcfce7')),
                ("Completion Rate", f"{metrics.get('completion_rate', 0)}%", colors.HexColor('#ede9fe')),
                ("Total Session Hours", str(metrics.get('avg_hours', 0)), colors.HexColor('#fef3c7')),
                ("Departments", str(df['Department'].nunique()) if 'Department' in df.columns else '0', colors.HexColor('#fee2e2')),
                ("Business Units", str(df['Business Unit'].nunique()) if 'Business Unit' in df.columns else '0', colors.HexColor('#e2e8f0')),
            ]

            cells = []
            for title_text, value_text, bg_color in card_items:
                card_content = Table([[
                    Paragraph(title_text, styles['CardLabel'])
                ], [
                    Paragraph(f"<b>{value_text}</b>", styles['CardValue'])
                ]], colWidths=[2.8 * inch])
                card_content.setStyle(TableStyle([
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                cells.append(card_content)

            table = Table([cells[:3], cells[3:]], colWidths=[3.22 * inch] * 3, rowHeights=[0.9 * inch, 0.9 * inch])
            table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.6, colors.HexColor('#cbd5e1')),
                ('BACKGROUND', (0, 0), (0, 0), card_items[0][2]),
                ('BACKGROUND', (1, 0), (1, 0), card_items[1][2]),
                ('BACKGROUND', (2, 0), (2, 0), card_items[2][2]),
                ('BACKGROUND', (0, 1), (0, 1), card_items[3][2]),
                ('BACKGROUND', (1, 1), (1, 1), card_items[4][2]),
                ('BACKGROUND', (2, 1), (2, 1), card_items[5][2]),
            ]))
            return table

        def build_filter_panel():
            content = Paragraph(
                (
                    f"<b>Filters Applied</b><br/>"
                    f"Department: {dept_filter}<br/>"
                    f"Business Unit: {bu_filter}<br/>"
                    f"Date Range: {start_date or 'N/A'} to {end_date or 'N/A'}"
                ),
                styles['Subtle'],
            )
            panel = Table([[content]], colWidths=[9.75 * inch])
            panel.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
                ('BOX', (0, 0), (-1, -1), 0.7, colors.HexColor('#cbd5e1')),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            return panel

        def build_horizontal_bar_chart(title_text, items, width=340, height=190, max_value=None, suffix=''):
            drawing = Drawing(width, height)
            drawing.add(String(0, height - 14, title_text, fontName='Helvetica-Bold', fontSize=11, fillColor=colors.HexColor('#0f172a')))

            if not items:
                drawing.add(String(0, height / 2, "No data available", fontSize=9, fillColor=colors.HexColor('#64748b')))
                return drawing

            left_pad = 120
            right_pad = 26
            top_pad = 28
            bottom_pad = 24
            usable_w = width - left_pad - right_pad
            usable_h = height - top_pad - bottom_pad
            max_val = max_value if max_value is not None else _nice_max(max(val for _, val in items), 100 if suffix == '%' else 10)

            for step in range(5):
                x = left_pad + (usable_w * step / 4)
                value_label = max_val * step / 4
                drawing.add(Line(x, bottom_pad, x, bottom_pad + usable_h, strokeColor=colors.HexColor('#e2e8f0'), strokeWidth=0.6))
                label_text = f"{int(round(value_label))}{suffix}"
                drawing.add(String(x - 6, 8, label_text, fontSize=6, fillColor=colors.HexColor('#64748b')))

            row_height = usable_h / max(len(items), 1)
            bar_height = min(14, row_height * 0.6)

            for idx, (label, value) in enumerate(items):
                y_center = bottom_pad + usable_h - ((idx + 0.5) * row_height)
                y = y_center - (bar_height / 2)
                bar_w = 0 if max_val == 0 else (value / max_val) * usable_w
                drawing.add(String(0, y_center - 3, _truncate(label, 22), fontSize=7, fillColor=colors.HexColor('#334155')))
                drawing.add(Rect(left_pad, y, usable_w, bar_height, fillColor=colors.HexColor('#f1f5f9'), strokeColor=colors.HexColor('#e2e8f0'), strokeWidth=0.4))
                drawing.add(Rect(left_pad, y, bar_w, bar_height, fillColor=colors.HexColor('#009ADA'), strokeColor=colors.HexColor('#009ADA'), strokeWidth=0.4))
                drawing.add(String(left_pad + min(bar_w + 4, usable_w - 6), y_center - 3, f"{value:.1f}{suffix}" if isinstance(value, float) else f"{value}{suffix}", fontSize=7, fillColor=colors.HexColor('#0f172a')))

            return drawing

        def build_vertical_bar_chart(title_text, items, width=340, height=190, max_value=None):
            drawing = Drawing(width, height)
            drawing.add(String(0, height - 14, title_text, fontName='Helvetica-Bold', fontSize=11, fillColor=colors.HexColor('#0f172a')))

            if not items:
                drawing.add(String(0, height / 2, "No data available", fontSize=9, fillColor=colors.HexColor('#64748b')))
                return drawing

            left_pad = 28
            right_pad = 12
            top_pad = 28
            bottom_pad = 38
            usable_w = width - left_pad - right_pad
            usable_h = height - top_pad - bottom_pad
            max_val = max_value if max_value is not None else _nice_max(max(val for _, val in items), 10)

            for step in range(5):
                y = bottom_pad + (usable_h * step / 4)
                value_label = max_val * step / 4
                drawing.add(Line(left_pad, y, left_pad + usable_w, y, strokeColor=colors.HexColor('#e2e8f0'), strokeWidth=0.6))
                drawing.add(String(2, y - 2, str(int(round(value_label))), fontSize=6, fillColor=colors.HexColor('#64748b')))

            slot_w = usable_w / max(len(items), 1)
            bar_w = min(26, slot_w * 0.55)

            for idx, (label, value) in enumerate(items):
                x = left_pad + (idx * slot_w) + ((slot_w - bar_w) / 2)
                bar_h = 0 if max_val == 0 else (value / max_val) * usable_h
                drawing.add(Rect(x, bottom_pad, bar_w, bar_h, fillColor=colors.HexColor('#0f766e'), strokeColor=colors.HexColor('#0f766e'), strokeWidth=0.4))
                drawing.add(String(x + 2, bottom_pad + bar_h + 4, str(int(value)), fontSize=7, fillColor=colors.HexColor('#0f172a')))
                drawing.add(String(x - 2, 12, _truncate(label, 10), fontSize=7, fillColor=colors.HexColor('#334155')))

            return drawing

        def build_hours_category_table(title_text, label_header, categories):
            heading = Paragraph(f"<b>{title_text}</b>", styles['SectionTitle'])
            note = Paragraph(
                "A (>=20), B (15-19), C (10-14), D (5-9), E (0.1-4), F (0)",
                styles['Subtle'],
            )
            rows = [[label_header, 'A', 'B', 'C', 'D', 'E', 'F', 'Total']]

            totals = {key: 0 for key in ['A', 'B', 'C', 'D', 'E', 'F', 'total']}
            for label, data in categories.items():
                rows.append([
                    str(label),
                    str(data.get('A', 0)),
                    str(data.get('B', 0)),
                    str(data.get('C', 0)),
                    str(data.get('D', 0)),
                    str(data.get('E', 0)),
                    str(data.get('F', 0)),
                    str(data.get('total', 0)),
                ])
                for key in totals:
                    totals[key] += int(data.get(key, 0))

            rows.append([
                'Grand Total',
                str(totals['A']),
                str(totals['B']),
                str(totals['C']),
                str(totals['D']),
                str(totals['E']),
                str(totals['F']),
                str(totals['total']),
            ])

            col_widths = [2.55 * inch] + [0.60 * inch] * 6 + [0.78 * inch]
            table = Table(rows, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -2), colors.white),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dbeafe')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1e3a8a')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))

            return KeepTogether([
                heading,
                Spacer(1, 0.05 * inch),
                note,
                Spacer(1, 0.1 * inch),
                table,
            ])

        # Create PDF buffer
        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            leftMargin=0.45 * inch,
            rightMargin=0.45 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )
        story = []
        
        # Add title
        title = Paragraph(f"<b>Training Analytics Report</b>", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.12 * inch))
        
        # Add generation date and filters
        gen_date = Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Subtle'])
        story.append(gen_date)
        story.append(Spacer(1, 0.16 * inch))
        story.append(build_filter_panel())
        story.append(Spacer(1, 0.18 * inch))
        story.append(Paragraph("<b>Snapshot Metrics</b>", styles['SectionTitle']))
        story.append(Spacer(1, 0.12 * inch))
        story.append(build_metric_cards())
        story.append(Spacer(1, 0.16 * inch))
        story.append(Paragraph("<b>Highlights</b>", styles['SectionTitle']))
        story.append(Spacer(1, 0.08 * inch))
        story.append(build_highlights_panel())

        story.append(PageBreak())
        story.append(Paragraph("<b>Visual Summary</b>", styles['SectionTitle']))
        story.append(Spacer(1, 0.1 * inch))

        dept_pairs = sorted(_series_pairs(metrics.get('dept')), key=lambda item: item[1], reverse=True)[:6]
        bu_pairs = sorted(_series_pairs(metrics.get('bu')), key=lambda item: item[1], reverse=True)
        ttype_pairs = sorted(_series_pairs(metrics.get('ttype')), key=lambda item: item[1], reverse=True)[:5]
        ttype_20_pairs = sorted(_series_pairs(metrics.get('ttype_20hrs')), key=lambda item: item[1], reverse=True)[:5]
        hour_totals = _hours_totals(metrics.get('dept_hours_categories', {}))
        hour_pairs = [(label, hour_totals[label]) for label in ['A', 'B', 'C', 'D', 'E', 'F']]

        charts_grid = Table([
            [
                build_horizontal_bar_chart("Top Department Completion %", dept_pairs, max_value=100, suffix='%'),
                build_horizontal_bar_chart("Business Unit Completion %", bu_pairs, max_value=100, suffix='%'),
            ],
            [
                build_vertical_bar_chart("Overall Hours Category Distribution", hour_pairs, max_value=_nice_max(max([value for _, value in hour_pairs], default=0), 10)),
                build_horizontal_bar_chart("20H+ Training Type Share", ttype_20_pairs, max_value=100, suffix='%'),
            ],
        ], colWidths=[4.92 * inch, 4.92 * inch])
        charts_grid.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(charts_grid)

        story.append(PageBreak())
        story.append(Paragraph("<b>Training Type Insights</b>", styles['SectionTitle']))
        story.append(Spacer(1, 0.08 * inch))
        story.append(Paragraph(
            "Comparison of overall training mix versus training types that reached 20+ planned hours.",
            styles['Subtle'],
        ))
        story.append(Spacer(1, 0.1 * inch))

        training_type_grid = Table([
            [
                build_horizontal_bar_chart("Training Type Share (All)", ttype_pairs, width=340, height=150, max_value=100, suffix='%'),
                build_horizontal_bar_chart("Training Type Share (20H+ Completion)", ttype_20_pairs, width=340, height=150, max_value=100, suffix='%'),
            ],
        ], colWidths=[4.92 * inch, 4.92 * inch])
        training_type_grid.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(training_type_grid)
        story.append(Spacer(1, 0.12 * inch))
        story.append(build_hours_category_table(
            "Business Unit Training Hours Categories",
            "Business Unit",
            metrics.get('bu_hours_categories', {}),
        ))

        story.append(PageBreak())
        story.append(build_hours_category_table(
            "Department-wise Training Hours Categories",
            "Department",
            metrics.get('dept_hours_categories', {}),
        ))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'analytics_export.pdf'
        )
    except ImportError:
        return jsonify({"error": "PDF export requires 'reportlab' package. Install with: pip install reportlab"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500
